import streamlit as st
import pandas as pd
from datetime import datetime
import io
import os
import openpyxl
from openpyxl.styles import Font, Alignment

# --- 1. 頁面基本設定 (保留原樣) ---
st.set_page_config(page_title="翌新空壓機報價系統", layout="wide")

st.markdown("""
    <style>
    .main { background-color: #f5f5f5; }
    .price-text { color: #E84118; font-size: 24px; font-weight: bold; }
    [data-testid="stVerticalBlock"] > div:has(div.stImage) {
        background-color: white; padding: 20px; border-radius: 15px;
        border: 1px solid #ddd; min-height: 450px; text-align: center;
    }
    </style>
    """, unsafe_allow_html=True)

# --- 2. 完整產品資料庫 (嚴格保留，絕不刪減) ---
if 'cart' not in st.session_state:
    st.session_state.cart = {} 

products = {
    "空壓機": [
        ("5馬空壓機", "air_5.png"), ("10馬永磁變頻空壓機HCV-10PM-A", "air_10.png"), ("20馬永磁變頻空壓機HCV-20PM-A", "air_20.png"), 
        ("30馬永磁變頻空壓機HCV-30PM-A", "air_30.png"), ("50馬永磁變頻空壓機HCV-50PM-A", "air_50.png"), ("750馬永磁變頻空壓機HCV-75PM-A", "air_75.png"), ("100馬永磁變頻空壓機HCV-100PM-A", "air_100.png")
    ],
    "儲氣筒": [
        ("105儲氣筒", "tank_105.png"), ("360儲氣筒", "tank_360.png"), ("660儲氣筒", "tank_660.png")
    ],
    "乾燥機": {
        "宙升": [
            ("5馬宙升乾燥機SD-005", "zs_dryer_5.png"), ("10馬宙升乾燥機SD-010	", "zs_dryer_10.png"), ("20馬宙升乾燥機SD-020", "zs_dryer_20.png"), 
            ("30馬宙升乾燥機SD-030", "zs_dryer_30.png"), ("50馬宙升乾燥機SD-050", "zs_dryer_50.png"), ("100馬宙升乾燥機", "zs_dryer_100.png")
        ],
        "艾冷": [
            ("5馬艾冷乾燥機", "al_dryer_5.png"), ("10馬艾冷乾燥機", "al_dryer_10.png"), ("20馬艾冷乾燥機", "al_dryer_20.png"), 
            ("30馬艾冷乾燥機", "al_dryer_30.png"), ("50馬艾冷乾燥機", "al_dryer_50.png"), ("100馬艾冷乾燥機", "al_dryer_100.png")
        ]
    },
    "選配配件": [
        ("Ckd自動排水器", "drainer_ckd.png"), ("電子式自動排水器", "drainer_e.png"), 
        ("前置旋風分離器", "separator.png"), ("超精密過濾器組", "filter.png"), ("超精密過濾器芯", "filter_core.png")
    ]
}

unit_map = {
    "105儲氣筒": "只", "360儲氣筒": "只", "660儲氣筒": "只",
    "Ckd自動排水器": "只", "電子式自動排水器": "只", "前置旋風分離器": "只",
    "超精密過濾器組": "只", "超精密過濾器芯": "只"
}

if 'price_config' not in st.session_state:
    st.session_state.price_config = {}
    for cat, items in products.items():
        if isinstance(items, dict):
            for sub in items.values():
                for name, _ in sub: st.session_state.price_config[name] = 0
        else:
            for name, _ in items: st.session_state.price_config[name] = 0

# --- 3. 側邊欄 (保留原樣) ---
st.sidebar.title("🏢 翌新後台管理")
customer_name = st.sidebar.text_input("客戶名稱", value="")
contact_person = st.sidebar.text_input("聯絡人", value="")

with st.sidebar.expander("⚙️ 價格調整"):
    for name in sorted(st.session_state.price_config.keys()):
        st.session_state.price_config[name] = st.sidebar.number_input(f"{name} 單價", value=st.session_state.price_config[name])

# --- 4. 主展示介面 (保留原樣) ---
st.title("請選擇設備類別")
tabs = st.tabs(["空壓機", "儲氣筒", "乾燥機", "選配配件"])

def display_items(item_list):
    cols = st.columns(3)
    for i, (name, img) in enumerate(item_list):
        with cols[i % 3]:
            if os.path.exists(img):
                st.image(img, width=250)
            st.write(f"**{name}**")
            if st.button(f"➕ 加入報價單", key=f"btn_{name}"):
                st.session_state.cart[name] = st.session_state.cart.get(name, 0) + 1
                st.toast(f"已加入: {name}")

with tabs[0]: display_items(products["空壓機"])
with tabs[1]: display_items(products["儲氣筒"])
with tabs[2]:
    brand = st.radio("選擇品牌", ["宙升", "艾冷"], horizontal=True)
    display_items(products["乾燥機"][brand])
with tabs[3]: display_items(products["選配配件"])

# --- 5. 報價清單與 EXCEL 輸出 (精確修改輸出位置) ---
st.divider()
if st.session_state.cart:
    st.subheader("📋 目前報價清單")
    table_data = []
    total_val = 0
    for name, qty in st.session_state.cart.items():
        p = st.session_state.price_config.get(name, 0)
        sub = p * qty
        total_val += sub
        table_data.append([name, unit_map.get(name, "台"), qty, f"${p:,}", f"${sub:,}"])
    
    st.table(pd.DataFrame(table_data, columns=["品名及規格", "單位", "數量", "單價", "金額"]))
    st.markdown(f"### <span class='price-text'>總計金額：${total_val:,}</span>", unsafe_allow_html=True)

    template_path = "翌新估價單EXCEL.xlsx"
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        bold_font = Font(name='新細明體', size=11, bold=True)
        right_align = Alignment(horizontal='right', vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')

        # 【修正 1】 H14將正確日期填入,H15顯示金額
        ws['H14'] = datetime.now().strftime('估價日期:%Y-%m-%d')
        ws['H14'].font = bold_font  # 套用加粗字體
        ws['H14'].alignment = Alignment(horizontal='left') # 設定靠左對齊
        # 1. 寫入文字內容
        ws['H15'] = "金額"
        ws['H15'] = ""      # 先徹底清空
        ws['H15'] = "金額"  # 再重新填入
        # 2. 設定為粗體 (沿用你程式中定義好的 bold_font)
        ws['H15'].font = bold_font
        # 3. 設定為靠左對齊
        ws['H15'].alignment = Alignment(horizontal='left', vertical='center')
        # 填入客戶資訊
        ws['B11'] = customer_name
        ws['B12'] = contact_person

        # 【修正 2】精確對照圖 2 欄位：NO(A), 品名(B), 數量(D), 單位(E), 金額(H)
        for i, (name, qty) in enumerate(st.session_state.cart.items()):
            row = 17 + i
            price = st.session_state.price_config.get(name, 0)
            
            # NO  (A欄 = 1)
            ws.cell(row=row, column=1, value=i+1).font = bold_font
            
            # 品名規格 (B欄 = 2)
            ws.cell(row=row, column=2, value=name).font = bold_font
            
            # 1. 單位 (D欄 = 4)
            c_unit = ws.cell(row=row, column=4, value=unit_map.get(name, "台"))
            c_unit.font = bold_font
            c_unit.alignment = center_align

            # 2. 數量 (E欄 = 5)
            c_qty = ws.cell(row=row, column=5, value=qty)
            c_qty.font = bold_font
            c_qty.alignment = center_align

            # 3. 金額 (I欄 = 9) <--- 建議改 9，才會對齊標題
            c_sub = ws.cell(row=row, column=9, value=price * qty)
            c_sub.font = bold_font
            c_sub.alignment = right_align

        # 【修正 3】總計位置 H36 靠右對齊
        ws['H36'] = total_val
        ws['H36'].font = Font(name='新細明體', size=12, bold=True)
        ws['H36'].alignment = right_align
        
        # 可選：合計字樣放在 G36
        ws['G36'] = "合計"
        ws['G36'].font = bold_font
        ws['G36'].alignment = right_align

        output = io.BytesIO()
        wb.save(output)
        st.download_button(label="📤 下載 翌新專業報價單 (Excel)", data=output.getvalue(), file_name=f"翌新報價_{customer_name}.xlsx")

    if st.button("🗑️ 清空重選"):
        st.session_state.cart = {}
        st.rerun()
